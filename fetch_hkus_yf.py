#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
fetch_hkus_yf.py —— 港股 / 美股 全市场日线下载（yfinance 版）
==============================================================

【为什么用这个】
A股 用 baostock（fetch_a_baostock.py）最稳；港股/美股 baostock 不覆盖，
这里用 yfinance（雅虎财经，开源）拉 OHLCV。代码列表来源全部为官方/免费源，
不依赖东方财富（东财接口容易被限频/断连）：
  - 美股：NASDAQ Trader 官方代码表（nasdaqlisted + otherlisted，免费，已剔除ETF/测试票）
  - 港股：港交所(HKEX)官方证券列表（免费 xlsx，~2700+ 只个股，已剔除权证/CBBC/债券/ETF）
          三层兜底：HKEX官方 → 东财clist → 内置蓝筹种子列表

【安装依赖】（一次性）
    pip3 install yfinance openpyxl --break-system-packages
    （openpyxl 用于解析港交所官方列表的 xlsx 文件）

【常用命令】
    python3 fetch_hkus_yf.py                      # 港股+美股
    python3 fetch_hkus_yf.py --markets HK         # 只下港股
    python3 fetch_hkus_yf.py --markets US --limit 50   # 美股前 50 只(调试)
    python3 fetch_hkus_yf.py --proxy http://127.0.0.1:7890   # 指定代理

【代理（国内必看）】
数据源是 Yahoo(query1.finance.yahoo.com)/NASDAQ/港交所，国内直连很慢、还会被
Yahoo 限流(Too Many Requests)。脚本会**自动读取环境变量代理**(https_proxy 等)，
也可用 --proxy 显式指定。注意：yfinance 的 curl_cffi 后端默认不认"系统代理"(Clash
的系统代理开关)，所以要么设终端环境变量、要么用 --proxy。
    - 自动(推荐)：先 `export https_proxy=http://127.0.0.1:7890 http_proxy=$https_proxy`
      再跑脚本；脚本会打印"使用代理: ..."确认。
    - 显式：`python3 fetch_hkus_yf.py --proxy http://127.0.0.1:7890`
    - 强制直连：`--proxy ''`
(端口 7890 换成你自己代理的端口)

【输出】data_HK.json / data_US.json，契约同 CONTRACT.md：
    kline = [date, open, high, low, close, volume]，升序。
"""

import os
import sys
import json
import time
import argparse
import datetime
import urllib.request

try:
    import yfinance as yf
except ImportError:
    print("缺少 yfinance，请先安装：pip3 install yfinance --break-system-packages")
    sys.exit(1)

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
UA = "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 Chrome/120 Safari/537.36"

# 代理：数据源(Yahoo/NASDAQ)在国内直连很慢/被限流，走代理快很多。
# 优先用 --proxy 指定，否则自动读环境变量(https_proxy/http_proxy 等)。
PROXY = None


def resolve_proxy(cli_proxy):
    """确定要用的代理：命令行 > 环境变量；都没有则 None。"""
    if cli_proxy:
        return cli_proxy
    for key in ("https_proxy", "HTTPS_PROXY", "http_proxy", "HTTP_PROXY",
                "all_proxy", "ALL_PROXY"):
        v = os.environ.get(key)
        if v:
            return v
    return None


def apply_proxy(proxy):
    """把代理应用到 yfinance(curl_cffi) + urllib(NASDAQ表) + curl(HKEX表)。"""
    global PROXY
    PROXY = proxy
    if not proxy:
        print("  未使用代理(直连)。若在国内抓 Yahoo 很慢或被限流，"
              "请加 --proxy http://127.0.0.1:7890 (换成你的代理端口)")
        return
    print("  使用代理:", proxy)
    # yfinance：curl_cffi 后端默认不读环境变量代理，必须显式设置
    try:
        yf.set_config(proxy=proxy, retries=5)
    except Exception as e:
        print("  yfinance 代理设置失败:", e)
    # urllib：装全局 opener，us_symbols 抓 NASDAQ 表时生效
    try:
        h = urllib.request.ProxyHandler({"http": proxy, "https": proxy})
        urllib.request.install_opener(urllib.request.build_opener(h))
    except Exception as e:
        print("  urllib 代理设置失败:", e)

# 港股取不到 clist 时的内置种子（部分蓝筹，保证脚本可用）
HK_SEED = ["00700", "00005", "09988", "00939", "01299", "00388", "03690",
           "01810", "02318", "00883", "00857", "01024", "09618", "02020"]

# 美股取不到 NASDAQ 官方表时的内置种子（主要个股，保证脚本可用）
US_SEED = [
    ("AAPL", "Apple"), ("MSFT", "Microsoft"), ("NVDA", "NVIDIA"),
    ("AMZN", "Amazon"), ("GOOGL", "Alphabet"), ("META", "Meta"),
    ("TSLA", "Tesla"), ("AMD", "AMD"), ("NFLX", "Netflix"),
    ("INTC", "Intel"), ("BABA", "Alibaba"), ("PDD", "PDD"),
    ("JPM", "JPMorgan"), ("KO", "Coca-Cola"), ("DIS", "Disney"),
    ("BA", "Boeing"), ("PFE", "Pfizer"), ("F", "Ford"),
    ("NIO", "NIO"), ("PLTR", "Palantir"),
]


# ----------------------------------------------------------------------------
# 代码列表
# ----------------------------------------------------------------------------
def us_symbols(include_etf=False):
    """
    从 NASDAQ Trader 官方文件取美股代码 -> [(ticker, name)]。
    默认剔除 ETF 和 Test Issue（测试票）：本战法针对个股涨停催化剂，
    ETF 是一篮子基金，不存在"个股涨停"逻辑，混入会产生噪音信号。
    列结构：
      nasdaqlisted.txt: Symbol|SecurityName|MarketCategory|TestIssue|FinancialStatus|RoundLot|ETF|NextShares
      otherlisted.txt : ACTSymbol|SecurityName|Exchange|CQSSymbol|ETF|RoundLot|TestIssue|NASDAQSymbol
    """
    out = []
    # (候选url列表, code列, name列, ETF列, TestIssue列)
    # HTTPS 站点(www.nasdaqtrader.com)有反爬墙，会给非浏览器请求返回 JS 验证页(解析0行)；
    # FTP 源(ftp.nasdaqtrader.com)是同一份官方文件且无反爬，作为自动兜底
    files = [
        (["https://www.nasdaqtrader.com/dynamic/SymDir/nasdaqlisted.txt",
          "ftp://ftp.nasdaqtrader.com/SymbolDirectory/nasdaqlisted.txt"], 0, 1, 6, 3),
        (["https://www.nasdaqtrader.com/dynamic/SymDir/otherlisted.txt",
          "ftp://ftp.nasdaqtrader.com/SymbolDirectory/otherlisted.txt"], 0, 1, 4, 6),
    ]
    etf_skipped = test_skipped = 0
    for urls, ci, ni, ei, ti in files:
        parsed = None
        # 重试覆盖两种失败：①请求异常(网络抖动) ②请求成功但内容是错误页(解析出0行，
        # 常见于反爬墙/代理返回 200 的报错页面)。HTTPS 和 FTP 两个源轮流各试两次
        for attempt, url in enumerate(list(urls) * 2):
            txt = None
            try:
                req = urllib.request.Request(url, headers={"User-Agent": UA})
                txt = urllib.request.urlopen(req, timeout=25).read().decode("latin-1")
            except Exception as e:
                print("  美股代码表获取失败(第%d次): %s %s" % (attempt + 1, url, e))
                time.sleep(1.5 * (attempt + 1))
                continue
            rows, etf_n, test_n = [], 0, 0
            for line in txt.strip().split("\n")[1:]:
                if line.startswith("File Creation Time"):
                    continue
                parts = line.split("|")
                if len(parts) <= max(ci, ni, ei, ti):
                    continue
                sym = parts[ci].strip()
                name = parts[ni].strip()
                if parts[ti].strip() == "Y":
                    test_n += 1
                    continue
                if not include_etf and parts[ei].strip() == "Y":
                    etf_n += 1
                    continue
                # 过滤含特殊符号的衍生票(权证/优先股/多类别股等)
                if not sym or any(c in sym for c in ".$^") or len(sym) > 6:
                    continue
                rows.append((sym, name))
            if rows:
                parsed = (rows, etf_n, test_n)
                break
            print("  美股代码表内容异常(第%d次，解析出0行，可能是代理/站点返回了错误页): %s"
                  % (attempt + 1, url))
            time.sleep(1.5 * (attempt + 1))
        if parsed:
            out.extend(parsed[0])
            etf_skipped += parsed[1]
            test_skipped += parsed[2]
    # 去重
    seen, uniq = set(), []
    for s, n in out:
        if s not in seen:
            seen.add(s); uniq.append((s, n))
    print("  美股代码表：个股 %d 只（已剔除 ETF %d 只、测试票 %d 只）"
          % (len(uniq), etf_skipped, test_skipped))
    if not uniq:
        print("  NASDAQ 代码表不可用，使用内置美股种子列表(%d 只)" % len(US_SEED))
        uniq = list(US_SEED)
    return uniq


HKEX_LIST_URL = "https://www.hkex.com.hk/eng/services/trading/securities/securitieslists/ListOfSecurities.xlsx"
# 只保留真正的个股，剔除权证/牛熊证/债券/ETF/REITs等
HKEX_EQUITY_SUBCATS = {"Equity Securities (Main Board)", "Equity Securities (GEM)"}


def hk_symbols_hkex():
    """
    从港交所(HKEX)官方证券列表取港股代码 -> [(code5, name)]。
    该文件用 urllib 直接请求会卡死(域名对 urllib 异常)，改用 curl 子进程下载；
    再用 openpyxl 解析（需要 pip3 install openpyxl）。
    注意：该 xlsx 的 <dimension> 标签是错的(虚报只有几行)，openpyxl 的
    read_only 模式会因此提前截断，必须用非 read_only 模式完整加载。
    失败返回 []，由调用方决定是否走其它兜底源。
    """
    import subprocess
    import tempfile
    try:
        import openpyxl
    except ImportError:
        print("  缺少 openpyxl（HKEX 官方列表解析需要），请先安装："
              "pip3 install openpyxl --break-system-packages")
        return []
    tmp_path = os.path.join(tempfile.gettempdir(), "hkex_list.xlsx")
    curl_cmd = ["curl", "-sL", "--connect-timeout", "15", "--max-time", "40"]
    if PROXY:
        curl_cmd += ["-x", PROXY]
    curl_cmd += ["-o", tmp_path, HKEX_LIST_URL]
    try:
        r = subprocess.run(curl_cmd, timeout=50)
        if r.returncode != 0 or not os.path.exists(tmp_path) or os.path.getsize(tmp_path) < 1000:
            print("  HKEX 官方列表下载失败(curl 退出码 %s)" % r.returncode)
            return []
    except Exception as e:
        print("  HKEX 官方列表下载异常: %s" % e)
        return []
    try:
        wb = openpyxl.load_workbook(tmp_path, read_only=False, data_only=True)
        ws = wb[wb.sheetnames[0]]
        out = []
        for row in ws.iter_rows(min_row=4, values_only=True):
            if not row or not row[0]:
                continue
            code, name, category, subcat = row[0], row[1], row[2], row[3]
            if category == "Equity" and subcat in HKEX_EQUITY_SUBCATS:
                out.append((str(code).strip(), str(name).strip() if name else str(code)))
        return out
    except Exception as e:
        print("  HKEX 官方列表解析失败: %s" % e)
        return []
    finally:
        try:
            os.remove(tmp_path)
        except OSError:
            pass


def hk_symbols_eastmoney():
    """从东财 clist 取港股代码 -> [(code5, name)]（HKEX 官方源失败时的第二兜底）。"""
    out = []
    fs = "m:116+t:1,m:116+t:2,m:116+t:3,m:116+t:4"
    pn = 1
    while True:
        url = ("https://push2.eastmoney.com/api/qt/clist/get?pn=%d&pz=100&po=1&np=1"
               "&fltt=2&invt=2&fid=f3&fs=%s&fields=f12,f14" % (pn, fs))
        try:
            req = urllib.request.Request(url, headers={"User-Agent": UA})
            data = json.loads(urllib.request.urlopen(req, timeout=15).read().decode())
        except Exception as e:
            print("  港股 clist 第 %d 页失败: %s" % (pn, e))
            break
        d = data.get("data")
        if not d or not d.get("diff"):
            break
        for x in d["diff"]:
            out.append((str(x["f12"]), str(x.get("f14", x["f12"]))))
        if len(d["diff"]) < 100 or len(out) >= d.get("total", 0):
            break
        pn += 1
        time.sleep(0.1)
    return out


def hk_symbols():
    """
    港股代码列表，三层兜底：
    ① 港交所官方列表(HKEX，最全最准，~2700+只) → ② 东财 clist → ③ 内置蓝筹种子。
    """
    out = hk_symbols_hkex()
    if out:
        print("  港股代码表：HKEX 官方列表，个股 %d 只" % len(out))
        return out
    print("  HKEX 官方列表不可用，尝试东财 clist ...")
    out = hk_symbols_eastmoney()
    if out:
        print("  港股代码表：东财 clist，%d 只" % len(out))
        return out
    print("  东财 clist 也不可用，使用内置种子列表(%d 只)" % len(HK_SEED))
    return [(c, c) for c in HK_SEED]


def hk_ticker(code5):
    """港股代码 -> 雅虎 ticker：去前导零补 4 位 + .HK。00700->0700.HK, 09988->9988.HK"""
    try:
        n = int(code5)
    except ValueError:
        return None
    return "%04d.HK" % n


# ----------------------------------------------------------------------------
# 下载
# ----------------------------------------------------------------------------
MAX_KLINE_KEEP = 500  # 单只股票本地最多保留的交易日数（增量模式下防止文件无限增长）


def df_to_kline(df):
    """yfinance 单只 DataFrame -> [[date,open,high,low,close,volume], ...] 升序。

    权证/停牌等无价行为 None/NaN，float()/int() 会抛异常被跳过。
    yfinance 港股源的 open 口径偶尔落在当日 high/low 区间外(未收盘bar同理)，
    这里强制 high=max(o,h,l,c)、low=min(o,h,l,c) 保证 OHLC 自洽——战法用
    high/low 取区间极值，按四价极值收敛是正确且安全的修正。
    """
    kline = []
    for ts, row in df.iterrows():
        try:
            o = float(row["Open"]); h = float(row["High"])
            lo = float(row["Low"]); c = float(row["Close"]); v = int(row["Volume"])
        except (ValueError, TypeError, KeyError):
            continue
        if any(x != x for x in (o, h, lo, c)):  # NaN 跳过
            continue
        hi = max(o, h, lo, c)
        low = min(o, h, lo, c)
        kline.append([str(ts.date()), round(o, 4), round(hi, 4),
                      round(low, 4), round(c, 4), v])
    kline.sort(key=lambda r: r[0])
    return kline


def merge_kline(old_kline, new_kline):
    """按日期去重合并两段 kline（new 覆盖同日期的 old），排序后裁剪到 MAX_KLINE_KEEP 条。"""
    by_date = {row[0]: row for row in old_kline}
    for row in new_kline:
        by_date[row[0]] = row
    merged = sorted(by_date.values(), key=lambda r: r[0])
    return merged[-MAX_KLINE_KEEP:]


def load_existing(market):
    """读取本地已有 data_<market>.json -> {code: stock_obj}；不存在/损坏则返回空字典。"""
    path = os.path.join(SCRIPT_DIR, "data_%s.json" % market)
    if not os.path.exists(path):
        return {}
    try:
        with open(path, encoding="utf-8") as f:
            payload = json.load(f)
        return {s["code"]: s for s in payload.get("stocks", [])}
    except (json.JSONDecodeError, KeyError, OSError):
        print("  本地 data_%s.json 读取失败，按无历史处理（将全量拉取）" % market)
        return {}


def load_existing_index(market):
    """读取本地已有 data_<market>.json 的 index 字段（大盘指数）；不存在则返回 None。"""
    path = os.path.join(SCRIPT_DIR, "data_%s.json" % market)
    if not os.path.exists(path):
        return None
    try:
        with open(path, encoding="utf-8") as f:
            return json.load(f).get("index")
    except (json.JSONDecodeError, KeyError, OSError):
        return None


# 大盘代理指数：目前只有美股引擎(平台/杯柄突破策略)的"大盘没有走弱"条件用得到
MARKET_INDEX_TICKER = {"US": "SPY"}


def fetch_market_index(market, period, incremental, existing_index):
    """拉取市场大盘指数(供美股突破策略的大盘过滤条件使用)，支持增量更新。"""
    ticker = MARKET_INDEX_TICKER.get(market)
    if not ticker:
        return None
    if incremental and existing_index and existing_index.get("kline"):
        last_date = existing_index["kline"][-1][0]
        # 从本地最后一天(而非次日)开始重抓：若上次是盘中跑的，最后一根是半天的
        # 不完整K线，重抓后被 merge_kline 的"新覆盖旧"替换成收盘完整K线
        incr_start = last_date
        kmap = download_batch([ticker], 1, start=incr_start)
        new_rows = kmap.get(ticker)
        kline = merge_kline(existing_index["kline"], new_rows) if new_rows else existing_index["kline"]
    else:
        kmap = download_batch([ticker], 1, period=period)
        kline = kmap.get(ticker) or (existing_index["kline"] if existing_index else None)
    if not kline:
        return existing_index
    return {"code": ticker, "name": ticker, "kline": kline}


def download_batch(tickers, batch, period=None, start=None):
    """
    分批调用 yf.download，返回 {ticker: kline}（下载失败/无数据的 ticker 不在结果里）。
    period 和 start 二选一：period 用于全量拉取，start（YYYY-MM-DD）用于增量拉取。
    """
    out = {}
    if not start and not period:
        # 既无 start 又无 period，无法确定下载区间，直接返回空。
        return out
    total_batches = (len(tickers) + batch - 1) // batch
    t_start = time.time()
    for bi, i in enumerate(range(0, len(tickers), batch)):
        chunk = tickers[i:i + batch]
        # 实时进度（flush 保证重定向到文件/管道时也能立刻看到）
        if bi == 0:
            print("  开始下载：共 %d 只，分 %d 批(每批 %d 只)"
                  % (len(tickers), total_batches, batch), flush=True)
        else:
            elapsed = time.time() - t_start
            eta = elapsed / i * (len(tickers) - i)
            print("  进度: 批 %d/%d（已完成 %d/%d 只，已用 %d 秒，预计还需 %d 秒）"
                  % (bi + 1, total_batches, i, len(tickers), elapsed, eta), flush=True)
        try:
            if start:
                data = yf.download(chunk, start=start, interval="1d",
                                   group_by="ticker", auto_adjust=False,
                                   progress=False, threads=True)
            else:
                data = yf.download(chunk, period=period, interval="1d",
                                   group_by="ticker", auto_adjust=False,
                                   progress=False, threads=True)
        except Exception as e:
            print("  批次 %d 下载异常: %s" % (i, e))
            continue
        if data is None or len(data) == 0:
            continue
        single_flat = (len(chunk) == 1 and getattr(data.columns, "nlevels", 1) == 1)
        for t in chunk:
            try:
                # group_by="ticker" 通常返回按 ticker 分组的 MultiIndex 列，
                # 即使 chunk 只有 1 只股票也一样，用 data[t] 取子表。
                # 少数 yfinance 版本在单只票时返回单层列(非 MultiIndex)，此时用整表兜底。
                df = data if single_flat else data[t]
            except (KeyError, TypeError):
                continue
            if df is None or len(df) == 0:
                continue
            df = df.dropna(how="all")
            kline = df_to_kline(df)
            if kline:
                out[t] = kline
    return out


def fetch_market(market, period="2y", limit=None, batch=40, incremental=False):
    if market == "US":
        print("获取美股代码表...")
        syms = us_symbols()
        board = "us"
        items = [(s, s, n) for s, n in syms]   # (code, yahoo_ticker, name)
    else:
        print("获取港股代码表...")
        syms = hk_symbols()
        board = "hk"
        items = []
        for code5, name in syms:
            t = hk_ticker(code5)
            if t:
                items.append((code5, t, name))

    if limit:
        items = items[:limit]
    print("待下载 %d 只" % len(items))

    existing = load_existing(market) if incremental else {}
    if incremental:
        print("  增量模式：本地已有 %d 只，只补最新数据" % len(existing))

    today = datetime.date.today().isoformat()
    results = []
    ok = uptodate = appended = fresh = 0

    if not incremental or not existing:
        # 全量模式：一次性按 period 拉取
        kmap = download_batch([it[1] for it in items], batch, period=period)
        for code, ticker, name in items:
            kline = kmap.get(ticker)
            if kline:
                results.append({"code": code, "name": name, "board": board,
                                "secid": ticker, "kline": kline})
                ok += 1; fresh += 1
    else:
        # 增量模式：按"本地最新日期"分组批量请求。注意 start=本地最后一天(而非次日)：
        # 上次若在盘中跑过，最后一根是半天的不完整K线，重抓这一天让 merge_kline 用
        # 收盘完整K线覆盖它；已收录完整的日子重抓一根也只是被同值覆盖，无副作用
        full_items, groups = [], {}
        carry = []
        for code, ticker, name in items:
            old = existing.get(code)
            if not old or not old.get("kline"):
                full_items.append((code, ticker, name))
                continue
            last_date = old["kline"][-1][0]
            if last_date > today:   # 只有异常的"未来日期"才直接沿用（正常到不了这个分支）
                carry.append((code, ticker, name, old))
                continue
            groups.setdefault(last_date, []).append((code, ticker, name, old))

        # 已最新，直接沿用本地数据
        for code, ticker, name, old in carry:
            results.append({"code": code, "name": name, "board": board,
                            "secid": ticker, "kline": old["kline"]})
            ok += 1; uptodate += 1

        # 新股/本地无历史：全量拉取
        if full_items:
            kmap = download_batch([it[1] for it in full_items], batch, period=period)
            for code, ticker, name in full_items:
                kline = kmap.get(ticker)
                if kline:
                    results.append({"code": code, "name": name, "board": board,
                                    "secid": ticker, "kline": kline})
                    ok += 1; fresh += 1

        # 按增量起始日分组批量拉取，与本地历史合并
        for incr_start, group in groups.items():
            kmap = download_batch([g[1] for g in group], batch, start=incr_start)
            for code, ticker, name, old in group:
                new_rows = kmap.get(ticker)
                kline = merge_kline(old["kline"], new_rows) if new_rows else old["kline"]
                results.append({"code": code, "name": name, "board": board,
                                "secid": ticker, "kline": kline})
                ok += 1; appended += 1

        print("  已最新 %d / 增量补数 %d / 全量首拉(新股或本地无历史) %d"
              % (uptodate, appended, fresh))

        # 数据安全网：本地已有、但本次代码表里没有的股票，原样保留，绝不丢弃。
        # 否则代码表源临时挂掉(如 NASDAQ 表抓不到退回20只种子)时，写盘会把
        # 几千只的全量数据覆盖成一小撮票 —— 增量更新只允许"没更新到"，不允许"丢数据"。
        seen = {s["code"] for s in results}
        kept = 0
        for code, old in existing.items():
            if code not in seen and old.get("kline"):
                results.append(old)
                kept += 1
        if kept:
            print("  警告：本次代码表未包含本地已有的 %d 只，已原样保留旧数据"
                  "（代码表源可能临时不可用，建议改天重跑增量刷新它们）" % kept)

    results.sort(key=lambda s: s["code"])
    payload = {"market": market,
               "generated_at": datetime.datetime.now().replace(microsecond=0).isoformat(),
               "count": len(results), "stocks": results}

    if market in MARKET_INDEX_TICKER:
        existing_index = load_existing_index(market) if incremental else None
        idx = fetch_market_index(market, period, incremental, existing_index)
        if idx:
            payload["index"] = idx
            print("  大盘指数(%s)：%d 个交易日，末日 %s"
                  % (idx["code"], len(idx["kline"]), idx["kline"][-1][0]))
        else:
            print("  大盘指数(%s)获取失败，美股引擎的'大盘过滤'条件将自动放行(不拦截)"
                  % MARKET_INDEX_TICKER[market])

    out_path = os.path.join(SCRIPT_DIR, "data_%s.json" % market)
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False)
    print("%s 完成：成功 %d / 共 %d，写入 %s" % (market, ok, len(items), out_path))
    return payload


def main():
    ap = argparse.ArgumentParser(description="yfinance 版 港股/美股 日线下载")
    ap.add_argument("--markets", default="HK,US", help="HK,US（默认两者）")
    ap.add_argument("--period", default="2y",
                    help="回溯时长，如 1y/2y/3y（默认 2y；低位窗口需≥250交易日，勿低于2y）")
    ap.add_argument("--limit", type=int, default=None, help="每市场只下前 N 只(调试)")
    ap.add_argument("--batch", type=int, default=40, help="yfinance 批量大小(默认 40)")
    ap.add_argument("--incremental", action="store_true",
                    help="增量更新：已收录股票只补最新数据，新股全量拉取（每天收盘后建议用这个）")
    ap.add_argument("--proxy", default=None,
                    help="代理地址，如 http://127.0.0.1:7890；不填则自动读环境变量代理。"
                         "国内抓 Yahoo 数据必须走代理，否则很慢/被限流。加 --proxy '' 可强制直连")
    args = ap.parse_args()

    # 代理：--proxy 优先；--proxy '' 强制直连；未指定则自动读环境变量
    proxy = args.proxy if args.proxy is not None else resolve_proxy(None)
    apply_proxy(proxy or None)

    for m in [x.strip().upper() for x in args.markets.split(",") if x.strip()]:
        if m in ("HK", "US"):
            fetch_market(m, period=args.period, limit=args.limit, batch=args.batch,
                        incremental=args.incremental)
        else:
            print("跳过未知市场:", m)


if __name__ == "__main__":
    main()
